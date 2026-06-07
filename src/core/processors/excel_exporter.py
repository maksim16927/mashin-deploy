"""
Модуль для экспорта результатов детекции в Excel
Создает Excel файл с координатами всех детектированных объектов
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import numpy as np

# Перевод классов знаков на русский
SIGN_NAMES_RU = {
    # Приоритет
    'prio_give_way':        'Уступите дорогу',
    'prio_stop':            'Стоп',
    'prio_priority_road':   'Главная дорога',
    # Запрещающие — скорость
    'forb_speed_over_5':    'Ограничение 5 км/ч',
    'forb_speed_over_10':   'Ограничение 10 км/ч',
    'forb_speed_over_20':   'Ограничение 20 км/ч',
    'forb_speed_over_30':   'Ограничение 30 км/ч',
    'forb_speed_over_40':   'Ограничение 40 км/ч',
    'forb_speed_over_50':   'Ограничение 50 км/ч',
    'forb_speed_over_60':   'Ограничение 60 км/ч',
    'forb_speed_over_70':   'Ограничение 70 км/ч',
    'forb_speed_over_80':   'Ограничение 80 км/ч',
    'forb_speed_over_90':   'Ограничение 90 км/ч',
    'forb_speed_over_100':  'Ограничение 100 км/ч',
    'forb_speed_over_110':  'Ограничение 110 км/ч',
    'forb_speed_over_120':  'Ограничение 120 км/ч',
    'forb_speed_over_130':  'Ограничение 130 км/ч',
    # Запрещающие — прочие
    'forb_no_entry':        'Въезд запрещён',
    'forb_no_parking':      'Стоянка запрещена',
    'forb_no_stopping':     'Остановка запрещена',
    'forb_overtake_car':    'Обгон запрещён',
    'forb_overtake_trucks': 'Обгон грузовикам запрещён',
    'forb_trucks':          'Движение грузовиков запрещено',
    'forb_turn_left':       'Поворот налево запрещён',
    'forb_turn_right':      'Поворот направо запрещён',
    'forb_weight_over_3.5t':'Ограничение массы 3.5т',
    'forb_weight_over_7.5t':'Ограничение массы 7.5т',
    'forb_u_turn':          'Разворот запрещён',
    # Информационные
    'info_bus_station':     'Автобусная остановка',
    'info_crosswalk':       'Пешеходный переход',
    'info_highway':         'Автомагистраль',
    'info_one_way':         'Одностороннее движение',
    'info_parking':         'Парковка',
    'info_taxi_parking':    'Парковка такси',
    # Предупреждающие
    'warn_children':        'Дети',
    'warn_construction':    'Дорожные работы',
    'warn_crosswalk':       'Пешеходный переход (пред.)',
    'warn_cyclists':        'Велосипедисты',
    'warn_left_curve':      'Поворот налево',
    'warn_right_curve':     'Поворот направо',
    'warn_domestic_animals':'Животные',
    'warn_other_dangers':   'Прочие опасности',
    'warn_poor_road_surface':'Неровная дорога',
    'warn_roundabout':      'Круговое движение',
    'warn_sharp_left_curve':'Крутой поворот налево',
    'warn_sharp_right_curve':'Крутой поворот направо',
    'warn_slippery_road':   'Скользкая дорога',
    'warn_hump':            'Лежачий полицейский',
    'warn_traffic_light':   'Светофор',
    'warn_tram':            'Трамвайная линия',
    'warn_two_way_traffic': 'Двустороннее движение',
    'warn_wild_animals':    'Дикие животные',
    # Предписывающие
    'mand_bike_lane':       'Велосипедная дорожка',
    'mand_go_left':         'Движение налево',
    'mand_go_left_right':   'Движение налево или направо',
    'mand_go_right':        'Движение направо',
    'mand_go_straight':     'Движение прямо',
    'mand_go_straight_left':'Движение прямо или налево',
    'mand_go_straight_right':'Движение прямо или направо',
    'mand_pass_left':       'Объезд справа',
    'mand_pass_left_right': 'Объезд с любой стороны',
    'mand_pass_right':      'Объезд слева',
    'mand_roundabout':      'Круговое движение (пред.)',
    # Общие / нераспознанные
    'speed_limit_20':  'Ограничение 20 км/ч',
    'speed_limit_30':  'Ограничение 30 км/ч',
    'speed_limit_40':  'Ограничение 40 км/ч',
    'speed_limit_50':  'Ограничение 50 км/ч',
    'speed_limit_60':  'Ограничение 60 км/ч',
    'speed_limit_70':  'Ограничение 70 км/ч',
    'speed_limit_80':  'Ограничение 80 км/ч',
    'speed_limit_90':  'Ограничение 90 км/ч',
    'speed_limit_100': 'Ограничение 100 км/ч',
    'speed_limit_120': 'Ограничение 120 км/ч',
    'stop':            'Стоп',
    'Pedestrian_Crossing': 'Пешеходный переход',
    'Parking-Sign':    'Парковка',
    'Round-About':     'Круговое движение',
    'Warning':         'Предупреждение',
    'bump':            'Лежачий полицейский',
    'do_not_enter':    'Въезд запрещён',
    'do_not_u_turn':   'Разворот запрещён',
    'no_parking':      'Стоянка запрещена',
    'no_waiting':      'Остановка запрещена',
    'u_turn':          'Разворот',
}

def get_sign_name_ru(class_name: str) -> str:
    """Возвращает русское название знака или форматирует техническое"""
    if class_name in SIGN_NAMES_RU:
        return SIGN_NAMES_RU[class_name]
    # Пробуем автоматически форматировать
    name = class_name.replace('_', ' ').replace('forb ', '').replace('warn ', '').replace('mand ', '').replace('info ', '').replace('prio ', '')
    return name.capitalize()


def _format_timestamp(seconds) -> str:
    """Форматирует секунды в MM:SS"""
    if seconds is None:
        return ''
    try:
        s = int(float(seconds))
        return f"{s // 60:02d}:{s % 60:02d}"
    except Exception:
        return ''


def create_excel_report(
    detections: List[Dict],
    output_path: Optional[str] = None,
    video_name: Optional[str] = None
) -> str:
    if not detections:
        raise ValueError("Список детекций пуст")

    # --- Строим общий DataFrame ---
    df_data = []
    for det in detections:
        bbox = det.get('bbox', (0, 0, 0, 0))
        coords = det.get('coordinates')
        picket = det.get('picket')
        ts = det.get('timestamp')
        class_name = det.get('class_name', '')
        obj_type = det.get('object_type', '')

        row = {
            'Тип объекта':  obj_type,
            'Класс (код)':  class_name,
            'Название':     get_sign_name_ru(class_name) if obj_type == 'sign' else class_name,
            'Уверенность %': round(float(det.get('confidence', 0)) * 100, 1),
            'Кадр':         det.get('frame_number', ''),
            'Время':        _format_timestamp(ts),
        }
        if picket is not None:
            km = int(picket) // 1000
            m  = int(picket) % 1000
            row['Пикетаж'] = f"КМ {km}+{m:03d}"
            row['Пикетаж (м)'] = int(picket)
        if coords and len(coords) >= 2:
            row['Широта']  = coords[0]
            row['Долгота'] = coords[1]
        row.update({'X1': bbox[0], 'Y1': bbox[1], 'X2': bbox[2], 'Y2': bbox[3]})
        df_data.append(row)

    df = pd.DataFrame(df_data)

    # --- Путь для сохранения ---
    if output_path is None:
        ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        video_prefix = video_name or "video"
        output_path = f"videos/excel_reports/{video_prefix}_{ts_str}_detections.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # Лист 1: Дорожные знаки (главный)
        df_signs = df[df['Тип объекта'] == 'sign'].copy()
        if not df_signs.empty:
            sign_cols = [c for c in ['Название', 'Класс (код)', 'Уверенность %', 'Время', 'Пикетаж', 'Пикетаж (м)', 'Широта', 'Долгота', 'Кадр'] if c in df_signs.columns]
            df_signs[sign_cols].to_excel(writer, sheet_name='Дорожные знаки', index=False)
            ws = writer.sheets['Дорожные знаки']
            col_widths = {'Название': 35, 'Класс (код)': 28, 'Уверенность %': 14, 'Время': 10,
                          'Пикетаж': 14, 'Пикетаж (м)': 14, 'Широта': 14, 'Долгота': 14, 'Кадр': 8}
            for i, col in enumerate(sign_cols, 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col, 14)
            # Жирный заголовок
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="1F4E79")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            # Чередующаяся заливка строк
            light_fill = PatternFill("solid", fgColor="DCE6F1")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = light_fill

        # Лист 2: Все детекции
        all_cols = [c for c in ['Тип объекта', 'Название', 'Класс (код)', 'Уверенность %', 'Время', 'Пикетаж', 'Широта', 'Долгота', 'Кадр'] if c in df.columns]
        df[all_cols].to_excel(writer, sheet_name='Все детекции', index=False)

        # Лист 3: Статистика
        signs_df   = df[df['Тип объекта'] == 'sign']
        barrier_df = df[df['Тип объекта'] == 'barrier']
        stats_rows = [
            ['Всего уникальных объектов',    len(df)],
            ['Дорожных знаков',              len(signs_df)],
            ['Барьеров/ограждений',          len(barrier_df)],
            ['Прочей инфраструктуры',        len(df) - len(signs_df) - len(barrier_df)],
        ]
        if not signs_df.empty and 'Название' in signs_df.columns:
            stats_rows.append(['--- Знаки по типам ---', ''])
            for name, cnt in signs_df['Название'].value_counts().items():
                stats_rows.append([name, cnt])
        df_stats = pd.DataFrame(stats_rows, columns=['Метрика', 'Значение'])
        df_stats.to_excel(writer, sheet_name='Статистика', index=False)
        ws_s = writer.sheets['Статистика']
        ws_s.column_dimensions['A'].width = 40
        ws_s.column_dimensions['B'].width = 14

    return str(output_path)


def format_detections_for_excel(
    frame_number: int,
    object_type: str,
    class_name: str,
    confidence: float,
    bbox: tuple,
    timestamp: Optional[float] = None,
    coordinates: Optional[tuple] = None,
    picket: Optional[float] = None
) -> Dict:
    """
    Форматирование детекции для экспорта в Excel
    
    Args:
        frame_number: Номер кадра
        object_type: Тип объекта (sign, barrier, car, etc.)
        class_name: Название класса
        confidence: Уверенность (0-1)
        bbox: Координаты bbox (x1, y1, x2, y2)
        timestamp: Время в секундах
        coordinates: GPS координаты (lat, lon)
        picket: Пикетаж в метрах
    
    Returns:
        Dict: Форматированная детекция для Excel
    """
    return {
        'frame_number': frame_number,
        'object_type': object_type,
        'class_name': class_name,
        'confidence': confidence,
        'bbox': bbox,
        'timestamp': timestamp,
        'coordinates': coordinates,
        'picket': picket
    }
